import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.cell.cell import MergedCell  # Import MergedCell

def generate_monthly_statement(month, year):
    # Define the base project directory
    project_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Define the file path based on the month and year
    file_path = os.path.join(project_dir, f'Files/Texts/statement-{month}-{year}.txt')
    
    # Check if the file exists
    if not os.path.exists(file_path):
        print(f"No statement file found for {month} {year}")
        return
    
    # Read the data from the text file
    data = []
    with open(file_path, 'r') as file:
        for line in file:
            if line.strip():  # Skip empty lines
                data.append(line.strip().split(','))

    # Define the column names
    columns = ["Date", "Invoice No", "Amount", "Vat(15%)", "Debit", "Credit", "Balance"]
    
    # Create a DataFrame from the data
    df = pd.DataFrame(data, columns=columns)
    
    # Convert the Date column to datetime format, handling errors
    df["Date"] = pd.to_datetime(df["Date"], format='%Y-%m-%d', errors='coerce')
    
    # Convert the relevant columns to float
    df["Amount"] = df["Amount"].astype(float)
    df["Vat(15%)"] = df["Vat(15%)"].astype(float)
    df["Debit"] = df["Debit"].astype(float)
    df["Credit"] = df["Credit"].astype(float)
    df["Balance"] = df["Balance"].astype(float)
    
    # Treat the Invoice No column as text
    df["Invoice No"] = df["Invoice No"].astype(str)
    
    # Drop rows with invalid dates
    df = df.dropna(subset=["Date"])
    
    # Drop duplicate invoices
    df = df.drop_duplicates(subset=["Invoice No"])
    
    # Sort the DataFrame by Date and Invoice No
    df = df.sort_values(by=["Date", "Invoice No"])
    
    # Format the Date column as a string to show only the date
    df["Date"] = df["Date"].dt.strftime('%Y-%m-%d')
    
    # Copy data from the "Amount" column to the "Debit" column
    df["Debit"] = df["Amount"]
    
    # Update the Balance column
    df["Balance"] = df["Debit"] + df["Vat(15%)"] - df["Credit"]
    
    # Define the output Excel file path
    output_file_path = os.path.join(project_dir, f'Files/Statements/statement-{month}-{year}.xlsx')
    
    # Ensure the directory for the output file exists
    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
    
    # Save the DataFrame to an Excel file
    df.to_excel(output_file_path, index=False)
    
    # Load the workbook and select the active worksheet
    wb = load_workbook(output_file_path)
    ws = wb.active
    
    # Add header information before the statement data
    ws.insert_rows(1, 6)  # Insert 6 rows at the top for the header
    ws.merge_cells('A1:G1')
    ws['A1'] = 'FAKHR ALTASHYEED EST'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Mobile: 00966 507006855'
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f'statement-{month}-{year}'
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A5'] = 'Vendor No: 337337'
    ws['E5'] = 'VAT NO : 310308065300003'
    
    # Merge cells in the Date column for rows with the same date and mention the number of invoices
    current_date = None
    start_row = None
    invoice_count = 0
    for row in range(7, ws.max_row + 1):  # Start from the seventh row (first six rows are the header)
        cell_value = ws[f'A{row}'].value
        if cell_value != current_date:
            if start_row is not None and start_row != row - 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                ws[f'A{start_row}'] = f"{current_date} ({invoice_count})"
                ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_date = cell_value
            start_row = row
            invoice_count = 1
        else:
            invoice_count += 1
    if start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
        ws[f'A{start_row}'] = f"{current_date} ({invoice_count})"
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths to show all data in cells without double-clicking
    for col in ws.iter_cols():
        max_length = 0
        column = None
        for cell in col:
            if not isinstance(cell, MergedCell) and cell.value is not None:
                if column is None:
                    column = cell.column_letter  # Get the column name
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        if column:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    # Add formulas to update other columns based on the amount
    for row in range(7, ws.max_row + 1):
        ws[f'D{row}'] = f'=C{row}*0.15'  # Vat is 15% of amount
        ws[f'E{row}'] = f'=C{row}'  # Debit is the same as amount
        ws[f'F{row}'] = 0  # Credit is set to 0
    
    # Add a delay to ensure VAT and Debit columns are updated before calculating Balance
    wb.save(output_file_path)
    wb = load_workbook(output_file_path)
    ws = wb.active
    
    # Update the Balance column after VAT and Debit columns are updated
    for row in range(7, ws.max_row + 1):
        ws[f'G{row}'] = f'=E{row}+D{row}-F{row}'  # Balance is debit + vat - credit
    
    # Add summary rows for totals
    add_summary_rows(ws)
    
    # Save the workbook
    wb.save(output_file_path)
    
    print(f"Monthly statement for {month} {year} has been generated: {output_file_path}")

def add_summary_rows(ws):
    max_row = ws.max_row
    total_invoices = max_row - 1  # Exclude the header row
    
    # Add summary rows with Excel formulas
    summary_data = [
        ("Total:", "", f"=SUM(C7:C{max_row})", f"=SUM(D7:D{max_row})", f"=SUM(E7:E{max_row})", "", f"=SUM(G7:G{max_row})"),
        ("Total no of invoices:", f"=COUNTA(B7:B{max_row})", "", "", "", "", ""),
        ("Total amount:", f"=SUM(C7:C{max_row})", "", "", "", "", ""),
        ("Total tax:", f"=SUM(D7:D{max_row})", "", "", "", "", ""),
        ("Total with tax:", f"=SUM(G7:G{max_row})", "", "", "", "", "")
    ]
    
    for summary_row in summary_data:
        ws.append(summary_row)
        for col in range(1, len(summary_row) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = Font(bold=True)

# Example usage
if __name__ == "__main__":
    # Get the current month and year
    current_date = datetime.now()
    last_month = current_date.replace(day=1) - pd.DateOffset(months=1)
    month = last_month.strftime("%B")
    year = last_month.year
    
    # Generate the monthly statement for the last month
    generate_monthly_statement(month, year)
